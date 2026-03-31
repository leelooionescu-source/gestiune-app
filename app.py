import os
import json
import tempfile
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, session
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from database import get_db, init_db
from models import User, create_admin_if_needed

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'gestiune-app-secret-key-2024')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max upload

UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Te rugăm să te autentifici.'


@login_manager.user_loader
def load_user(user_id):
    return User.get_by_id(int(user_id))


# ==================== AUTENTIFICARE ====================

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        user = User.get_by_username(username)
        if user and user.verify_password(password):
            login_user(user)
            return redirect(url_for('dashboard'))
        flash('Utilizator sau parolă incorectă.', 'danger')
    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))


@app.route('/utilizatori')
@login_required
def utilizatori_lista():
    if not current_user.is_admin:
        flash('Nu ai permisiunea de a accesa această pagină.', 'danger')
        return redirect(url_for('dashboard'))
    users = User.get_all()
    return render_template('utilizatori.html', utilizatori=users)


@app.route('/utilizatori/adauga', methods=['POST'])
@login_required
def utilizatori_adauga():
    if not current_user.is_admin:
        flash('Nu ai permisiunea.', 'danger')
        return redirect(url_for('dashboard'))
    username = request.form.get('username', '').strip()
    password = request.form.get('password', '')
    nume = request.form.get('nume', '').strip()
    is_admin = 1 if request.form.get('is_admin') else 0
    if not username or not password or not nume:
        flash('Toate câmpurile sunt obligatorii.', 'danger')
        return redirect(url_for('utilizatori_lista'))
    if User.get_by_username(username):
        flash('Acest utilizator există deja.', 'danger')
        return redirect(url_for('utilizatori_lista'))
    User.create(username, password, nume, is_admin)
    flash(f'Utilizatorul {nume} a fost adăugat.', 'success')
    return redirect(url_for('utilizatori_lista'))


@app.route('/utilizatori/sterge/<int:id>')
@login_required
def utilizatori_sterge(id):
    if not current_user.is_admin:
        flash('Nu ai permisiunea.', 'danger')
        return redirect(url_for('dashboard'))
    if id == current_user.id:
        flash('Nu te poți șterge pe tine.', 'danger')
        return redirect(url_for('utilizatori_lista'))
    User.delete(id)
    flash('Utilizator șters.', 'success')
    return redirect(url_for('utilizatori_lista'))


# ==================== DASHBOARD ====================

@app.route('/')
@login_required
def dashboard():
    db = get_db()
    stats = {
        'clienti': db.execute('SELECT COUNT(*) as cnt FROM clienti').fetchone()['cnt'],
        'contracte_active': db.execute("SELECT COUNT(*) as cnt FROM contracte WHERE status = 'Activ'").fetchone()['cnt'],
        'proiecte_in_lucru': db.execute("SELECT COUNT(*) as cnt FROM proiecte WHERE status = 'In lucru'").fetchone()['cnt'],
        'facturi_neincasate': db.execute("SELECT COUNT(*) as cnt FROM facturi WHERE status = 'Emisa'").fetchone()['cnt'],
        'valoare_neincasata': db.execute("SELECT COALESCE(SUM(valoare), 0) as val FROM facturi WHERE status = 'Emisa'").fetchone()['val'],
        'servicii_nepredate': db.execute("SELECT COUNT(*) as cnt FROM servicii WHERE status_predare = 'Nepredat'").fetchone()['cnt'],
        'servicii_nefacturate': db.execute("SELECT COUNT(*) as cnt FROM servicii WHERE status_facturare = 'Nefacturat'").fetchone()['cnt'],
        'servicii_neincasate': db.execute("SELECT COUNT(*) as cnt FROM servicii WHERE status_incasare = 'Neincasat'").fetchone()['cnt'],
        'valoare_servicii_neincasate': db.execute("SELECT COALESCE(SUM(valoare_totala), 0) as val FROM servicii WHERE status_incasare = 'Neincasat'").fetchone()['val'],
    }
    contracte_expira = db.execute("""
        SELECT c.*, cl.nume as client_nume
        FROM contracte c JOIN clienti cl ON c.client_id = cl.id
        WHERE c.status = 'Activ' AND c.data_sfarsit IS NOT NULL
        AND c.data_sfarsit <= date('now', '+30 days')
        ORDER BY c.data_sfarsit
        LIMIT 5
    """).fetchall()
    ultimele_facturi = db.execute("""
        SELECT f.*, ct.numar_contract
        FROM facturi f JOIN contracte ct ON f.contract_id = ct.id
        ORDER BY f.creat_la DESC LIMIT 5
    """).fetchall()
    db.close()
    return render_template('dashboard.html', stats=stats, contracte_expira=contracte_expira, ultimele_facturi=ultimele_facturi)


# ==================== CLIENTI ====================

@app.route('/clienti')
@login_required
def clienti_lista():
    db = get_db()
    cautare = request.args.get('cautare', '').strip()
    if cautare:
        clienti = db.execute(
            "SELECT * FROM clienti WHERE nume LIKE ? OR telefon LIKE ? OR email LIKE ? ORDER BY nume",
            (f'%{cautare}%', f'%{cautare}%', f'%{cautare}%')
        ).fetchall()
    else:
        clienti = db.execute('SELECT * FROM clienti ORDER BY nume').fetchall()
    db.close()
    return render_template('clienti/lista.html', clienti=clienti, cautare=cautare)


@app.route('/clienti/adauga', methods=['GET', 'POST'])
@login_required
def clienti_adauga():
    if request.method == 'POST':
        db = get_db()
        db.execute('INSERT INTO clienti (nume, telefon, email, notite) VALUES (?, ?, ?, ?)',
                   (request.form['nume'].strip(), request.form.get('telefon', '').strip(),
                    request.form.get('email', '').strip(), request.form.get('notite', '').strip()))
        db.commit()
        db.close()
        flash('Client adăugat cu succes.', 'success')
        return redirect(url_for('clienti_lista'))
    return render_template('clienti/formular.html', client=None)


@app.route('/clienti/editeaza/<int:id>', methods=['GET', 'POST'])
@login_required
def clienti_editeaza(id):
    db = get_db()
    if request.method == 'POST':
        db.execute('UPDATE clienti SET nume=?, telefon=?, email=?, notite=?, actualizat_la=CURRENT_TIMESTAMP WHERE id=?',
                   (request.form['nume'].strip(), request.form.get('telefon', '').strip(),
                    request.form.get('email', '').strip(), request.form.get('notite', '').strip(), id))
        db.commit()
        db.close()
        flash('Client actualizat.', 'success')
        return redirect(url_for('clienti_lista'))
    client = db.execute('SELECT * FROM clienti WHERE id = ?', (id,)).fetchone()
    db.close()
    if not client:
        flash('Client negăsit.', 'danger')
        return redirect(url_for('clienti_lista'))
    return render_template('clienti/formular.html', client=client)


@app.route('/clienti/sterge/<int:id>')
@login_required
def clienti_sterge(id):
    db = get_db()
    contracte = db.execute('SELECT COUNT(*) as cnt FROM contracte WHERE client_id = ?', (id,)).fetchone()['cnt']
    if contracte > 0:
        flash('Nu poți șterge clientul - are contracte asociate.', 'danger')
    else:
        db.execute('DELETE FROM clienti WHERE id = ?', (id,))
        db.commit()
        flash('Client șters.', 'success')
    db.close()
    return redirect(url_for('clienti_lista'))


@app.route('/clienti/<int:id>/detalii')
@login_required
def clienti_detalii(id):
    db = get_db()
    client = db.execute('SELECT * FROM clienti WHERE id = ?', (id,)).fetchone()
    if not client:
        db.close()
        flash('Client negăsit.', 'danger')
        return redirect(url_for('clienti_lista'))

    contracte = db.execute('SELECT * FROM contracte WHERE client_id = ? ORDER BY creat_la DESC', (id,)).fetchall()

    # Build contract data with HG-uri and servicii
    contracte_data = []
    for c in contracte:
        hg_uri = db.execute('SELECT * FROM hg_uri WHERE contract_id = ? ORDER BY creat_la', (c['id'],)).fetchall()
        servicii = db.execute('SELECT * FROM servicii WHERE contract_id = ? ORDER BY creat_la', (c['id'],)).fetchall()
        contracte_data.append({
            'contract': c,
            'hg_uri': hg_uri,
            'servicii': servicii
        })

    # Compute totals across all services for this client
    totals = db.execute("""
        SELECT
            COALESCE(SUM(s.numar_imobile), 0) as total_imobile,
            COALESCE(SUM(s.valoare_totala), 0) as total_valoare,
            SUM(CASE WHEN s.status_predare = 'Predat' THEN 1 ELSE 0 END) as predate,
            SUM(CASE WHEN s.status_predare = 'Nepredat' THEN 1 ELSE 0 END) as nepredate,
            SUM(CASE WHEN s.status_facturare = 'Facturat' THEN 1 ELSE 0 END) as facturate,
            SUM(CASE WHEN s.status_facturare = 'Nefacturat' THEN 1 ELSE 0 END) as nefacturate,
            SUM(CASE WHEN s.status_incasare = 'Incasat' THEN 1 ELSE 0 END) as incasate,
            SUM(CASE WHEN s.status_incasare = 'Neincasat' THEN 1 ELSE 0 END) as neincasate,
            COALESCE(SUM(CASE WHEN s.status_incasare = 'Neincasat' THEN s.valoare_totala ELSE 0 END), 0) as valoare_neincasata
        FROM servicii s
        JOIN contracte c ON s.contract_id = c.id
        WHERE c.client_id = ?
    """, (id,)).fetchone()

    db.close()
    return render_template('clienti/detalii.html', client=client, contracte_data=contracte_data, totals=totals)


# ==================== HG-URI ====================

@app.route('/hg/adauga', methods=['POST'])
@login_required
def hg_adauga():
    contract_id = request.form['contract_id']
    client_id = request.form['client_id']
    db = get_db()
    db.execute('INSERT INTO hg_uri (contract_id, numar_hg, descriere) VALUES (?, ?, ?)',
               (contract_id, request.form['numar_hg'].strip(), request.form.get('descriere', '').strip()))
    db.commit()
    db.close()
    flash('HG adăugat cu succes.', 'success')
    return redirect(url_for('clienti_detalii', id=client_id))


@app.route('/hg/sterge/<int:id>')
@login_required
def hg_sterge(id):
    db = get_db()
    hg = db.execute("""SELECT h.*, c.client_id FROM hg_uri h
                       JOIN contracte c ON h.contract_id = c.id WHERE h.id = ?""", (id,)).fetchone()
    if hg:
        client_id = hg['client_id']
        db.execute('DELETE FROM hg_uri WHERE id = ?', (id,))
        db.commit()
        flash('HG șters.', 'success')
    else:
        client_id = None
        flash('HG negăsit.', 'danger')
    db.close()
    if client_id:
        return redirect(url_for('clienti_detalii', id=client_id))
    return redirect(url_for('clienti_lista'))


# ==================== SERVICII ====================

@app.route('/servicii/adauga', methods=['POST'])
@login_required
def servicii_adauga():
    contract_id = request.form['contract_id']
    client_id = request.form['client_id']
    numar_imobile = int(request.form.get('numar_imobile', 0) or 0)
    pret_per_imobil = float(request.form.get('pret_per_imobil', 0) or 0)
    valoare_totala = numar_imobile * pret_per_imobil
    db = get_db()
    db.execute('''INSERT INTO servicii (contract_id, descriere_serviciu, numar_imobile, pret_per_imobil,
                  valoare_totala, status_predare, status_facturare, status_incasare)
                  VALUES (?, ?, ?, ?, ?, 'Nepredat', 'Nefacturat', 'Neincasat')''',
               (contract_id, request.form['descriere_serviciu'].strip(),
                numar_imobile, pret_per_imobil, valoare_totala))
    db.commit()
    db.close()
    flash('Serviciu adăugat cu succes.', 'success')
    return redirect(url_for('clienti_detalii', id=client_id))


@app.route('/servicii/editeaza/<int:id>', methods=['POST'])
@login_required
def servicii_editeaza(id):
    client_id = request.form['client_id']
    numar_imobile = int(request.form.get('numar_imobile', 0) or 0)
    pret_per_imobil = float(request.form.get('pret_per_imobil', 0) or 0)
    valoare_totala = numar_imobile * pret_per_imobil
    db = get_db()
    db.execute('''UPDATE servicii SET descriere_serviciu=?, numar_imobile=?, pret_per_imobil=?,
                  valoare_totala=?, data_predare=?, status_predare=?, numar_factura=?, data_factura=?,
                  status_facturare=?, data_incasare=?, status_incasare=?, observatii=?,
                  actualizat_la=CURRENT_TIMESTAMP WHERE id=?''',
               (request.form.get('descriere_serviciu', '').strip(),
                numar_imobile, pret_per_imobil, valoare_totala,
                request.form.get('data_predare') or None, request.form.get('status_predare', 'Nepredat'),
                request.form.get('numar_factura', '').strip() or None,
                request.form.get('data_factura') or None, request.form.get('status_facturare', 'Nefacturat'),
                request.form.get('data_incasare') or None, request.form.get('status_incasare', 'Neincasat'),
                request.form.get('observatii', '').strip(), id))
    db.commit()
    db.close()
    flash('Serviciu actualizat.', 'success')
    return redirect(url_for('clienti_detalii', id=client_id))


@app.route('/servicii/sterge/<int:id>')
@login_required
def servicii_sterge(id):
    db = get_db()
    srv = db.execute("""SELECT s.*, c.client_id FROM servicii s
                        JOIN contracte c ON s.contract_id = c.id WHERE s.id = ?""", (id,)).fetchone()
    if srv:
        client_id = srv['client_id']
        db.execute('DELETE FROM servicii WHERE id = ?', (id,))
        db.commit()
        flash('Serviciu șters.', 'success')
    else:
        client_id = None
        flash('Serviciu negăsit.', 'danger')
    db.close()
    if client_id:
        return redirect(url_for('clienti_detalii', id=client_id))
    return redirect(url_for('clienti_lista'))


# ==================== CONTRACTE ====================

@app.route('/contracte')
@login_required
def contracte_lista():
    db = get_db()
    cautare = request.args.get('cautare', '').strip()
    status_filtru = request.args.get('status', '').strip()
    query = """SELECT c.*, cl.nume as client_nume
               FROM contracte c JOIN clienti cl ON c.client_id = cl.id WHERE 1=1"""
    params = []
    if cautare:
        query += " AND (c.numar_contract LIKE ? OR c.descriere LIKE ? OR cl.nume LIKE ?)"
        params.extend([f'%{cautare}%'] * 3)
    if status_filtru:
        query += " AND c.status = ?"
        params.append(status_filtru)
    query += " ORDER BY c.creat_la DESC"
    contracte = db.execute(query, params).fetchall()
    db.close()
    return render_template('contracte/lista.html', contracte=contracte, cautare=cautare, status_filtru=status_filtru)


@app.route('/contracte/adauga', methods=['GET', 'POST'])
@login_required
def contracte_adauga():
    db = get_db()
    if request.method == 'POST':
        db.execute('''INSERT INTO contracte (numar_contract, client_id, descriere, valoare, data_inceput, data_sfarsit, status)
                      VALUES (?, ?, ?, ?, ?, ?, ?)''',
                   (request.form['numar_contract'].strip(), request.form['client_id'],
                    request.form.get('descriere', '').strip(),
                    float(request.form['valoare']) if request.form.get('valoare') else None,
                    request.form.get('data_inceput') or None, request.form.get('data_sfarsit') or None,
                    request.form.get('status', 'Activ')))
        db.commit()
        db.close()
        flash('Contract adăugat cu succes.', 'success')
        return redirect(url_for('contracte_lista'))
    clienti = db.execute('SELECT id, nume FROM clienti ORDER BY nume').fetchall()
    db.close()
    return render_template('contracte/formular.html', contract=None, clienti=clienti)


@app.route('/contracte/editeaza/<int:id>', methods=['GET', 'POST'])
@login_required
def contracte_editeaza(id):
    db = get_db()
    if request.method == 'POST':
        db.execute('''UPDATE contracte SET numar_contract=?, client_id=?, descriere=?, valoare=?,
                      data_inceput=?, data_sfarsit=?, status=?, actualizat_la=CURRENT_TIMESTAMP WHERE id=?''',
                   (request.form['numar_contract'].strip(), request.form['client_id'],
                    request.form.get('descriere', '').strip(),
                    float(request.form['valoare']) if request.form.get('valoare') else None,
                    request.form.get('data_inceput') or None, request.form.get('data_sfarsit') or None,
                    request.form.get('status', 'Activ'), id))
        db.commit()
        db.close()
        flash('Contract actualizat.', 'success')
        return redirect(url_for('contracte_lista'))
    contract = db.execute('SELECT * FROM contracte WHERE id = ?', (id,)).fetchone()
    clienti = db.execute('SELECT id, nume FROM clienti ORDER BY nume').fetchall()
    db.close()
    if not contract:
        flash('Contract negăsit.', 'danger')
        return redirect(url_for('contracte_lista'))
    return render_template('contracte/formular.html', contract=contract, clienti=clienti)


@app.route('/contracte/sterge/<int:id>')
@login_required
def contracte_sterge(id):
    db = get_db()
    proiecte = db.execute('SELECT COUNT(*) as cnt FROM proiecte WHERE contract_id = ?', (id,)).fetchone()['cnt']
    facturi = db.execute('SELECT COUNT(*) as cnt FROM facturi WHERE contract_id = ?', (id,)).fetchone()['cnt']
    servicii = db.execute('SELECT COUNT(*) as cnt FROM servicii WHERE contract_id = ?', (id,)).fetchone()['cnt']
    hg_cnt = db.execute('SELECT COUNT(*) as cnt FROM hg_uri WHERE contract_id = ?', (id,)).fetchone()['cnt']
    if proiecte > 0 or facturi > 0 or servicii > 0 or hg_cnt > 0:
        flash('Nu poți șterge contractul - are proiecte, facturi, servicii sau HG-uri asociate.', 'danger')
    else:
        db.execute('DELETE FROM contracte WHERE id = ?', (id,))
        db.commit()
        flash('Contract șters.', 'success')
    db.close()
    return redirect(url_for('contracte_lista'))


# ==================== PROIECTE ====================

@app.route('/proiecte')
@login_required
def proiecte_lista():
    db = get_db()
    cautare = request.args.get('cautare', '').strip()
    status_filtru = request.args.get('status', '').strip()
    query = """SELECT p.*, c.numar_contract, cl.nume as client_nume
               FROM proiecte p
               JOIN contracte c ON p.contract_id = c.id
               JOIN clienti cl ON c.client_id = cl.id WHERE 1=1"""
    params = []
    if cautare:
        query += " AND (p.nume LIKE ? OR p.descriere LIKE ? OR p.responsabil LIKE ?)"
        params.extend([f'%{cautare}%'] * 3)
    if status_filtru:
        query += " AND p.status = ?"
        params.append(status_filtru)
    query += " ORDER BY p.creat_la DESC"
    proiecte = db.execute(query, params).fetchall()
    db.close()
    return render_template('proiecte/lista.html', proiecte=proiecte, cautare=cautare, status_filtru=status_filtru)


@app.route('/proiecte/adauga', methods=['GET', 'POST'])
@login_required
def proiecte_adauga():
    db = get_db()
    if request.method == 'POST':
        db.execute('''INSERT INTO proiecte (nume, contract_id, descriere, responsabil, data_start, data_estimata_finalizare, status)
                      VALUES (?, ?, ?, ?, ?, ?, ?)''',
                   (request.form['nume'].strip(), request.form['contract_id'],
                    request.form.get('descriere', '').strip(), request.form.get('responsabil', '').strip(),
                    request.form.get('data_start') or None, request.form.get('data_estimata_finalizare') or None,
                    request.form.get('status', 'In lucru')))
        db.commit()
        db.close()
        flash('Proiect adăugat cu succes.', 'success')
        return redirect(url_for('proiecte_lista'))
    contracte = db.execute("""SELECT c.id, c.numar_contract, cl.nume as client_nume
                              FROM contracte c JOIN clienti cl ON c.client_id = cl.id
                              ORDER BY c.numar_contract""").fetchall()
    db.close()
    return render_template('proiecte/formular.html', proiect=None, contracte=contracte)


@app.route('/proiecte/editeaza/<int:id>', methods=['GET', 'POST'])
@login_required
def proiecte_editeaza(id):
    db = get_db()
    if request.method == 'POST':
        db.execute('''UPDATE proiecte SET nume=?, contract_id=?, descriere=?, responsabil=?,
                      data_start=?, data_estimata_finalizare=?, status=?, actualizat_la=CURRENT_TIMESTAMP WHERE id=?''',
                   (request.form['nume'].strip(), request.form['contract_id'],
                    request.form.get('descriere', '').strip(), request.form.get('responsabil', '').strip(),
                    request.form.get('data_start') or None, request.form.get('data_estimata_finalizare') or None,
                    request.form.get('status', 'In lucru'), id))
        db.commit()
        db.close()
        flash('Proiect actualizat.', 'success')
        return redirect(url_for('proiecte_lista'))
    proiect = db.execute('SELECT * FROM proiecte WHERE id = ?', (id,)).fetchone()
    contracte = db.execute("""SELECT c.id, c.numar_contract, cl.nume as client_nume
                              FROM contracte c JOIN clienti cl ON c.client_id = cl.id
                              ORDER BY c.numar_contract""").fetchall()
    db.close()
    if not proiect:
        flash('Proiect negăsit.', 'danger')
        return redirect(url_for('proiecte_lista'))
    return render_template('proiecte/formular.html', proiect=proiect, contracte=contracte)


@app.route('/proiecte/sterge/<int:id>')
@login_required
def proiecte_sterge(id):
    db = get_db()
    predari = db.execute('SELECT COUNT(*) as cnt FROM predari WHERE proiect_id = ?', (id,)).fetchone()['cnt']
    if predari > 0:
        flash('Nu poți șterge proiectul - are predări asociate.', 'danger')
    else:
        db.execute('DELETE FROM proiecte WHERE id = ?', (id,))
        db.commit()
        flash('Proiect șters.', 'success')
    db.close()
    return redirect(url_for('proiecte_lista'))


# ==================== PREDARI ====================

@app.route('/predari')
@login_required
def predari_lista():
    db = get_db()
    cautare = request.args.get('cautare', '').strip()
    query = """SELECT pr.*, p.nume as proiect_nume, c.numar_contract, cl.nume as client_nume
               FROM predari pr
               JOIN proiecte p ON pr.proiect_id = p.id
               JOIN contracte c ON p.contract_id = c.id
               JOIN clienti cl ON c.client_id = cl.id WHERE 1=1"""
    params = []
    if cautare:
        query += " AND (pr.descriere LIKE ? OR p.nume LIKE ? OR cl.nume LIKE ?)"
        params.extend([f'%{cautare}%'] * 3)
    query += " ORDER BY pr.data_predare DESC"
    predari = db.execute(query, params).fetchall()
    db.close()
    return render_template('predari/lista.html', predari=predari, cautare=cautare)


@app.route('/predari/adauga', methods=['GET', 'POST'])
@login_required
def predari_adauga():
    db = get_db()
    if request.method == 'POST':
        db.execute('''INSERT INTO predari (proiect_id, data_predare, descriere, document_predare, observatii)
                      VALUES (?, ?, ?, ?, ?)''',
                   (request.form['proiect_id'], request.form['data_predare'],
                    request.form.get('descriere', '').strip(), request.form.get('document_predare', '').strip(),
                    request.form.get('observatii', '').strip()))
        db.commit()
        db.close()
        flash('Predare adăugată cu succes.', 'success')
        return redirect(url_for('predari_lista'))
    proiecte = db.execute("""SELECT p.id, p.nume, c.numar_contract, cl.nume as client_nume
                             FROM proiecte p
                             JOIN contracte c ON p.contract_id = c.id
                             JOIN clienti cl ON c.client_id = cl.id
                             ORDER BY p.nume""").fetchall()
    db.close()
    return render_template('predari/formular.html', predare=None, proiecte=proiecte)


@app.route('/predari/editeaza/<int:id>', methods=['GET', 'POST'])
@login_required
def predari_editeaza(id):
    db = get_db()
    if request.method == 'POST':
        db.execute('''UPDATE predari SET proiect_id=?, data_predare=?, descriere=?, document_predare=?, observatii=? WHERE id=?''',
                   (request.form['proiect_id'], request.form['data_predare'],
                    request.form.get('descriere', '').strip(), request.form.get('document_predare', '').strip(),
                    request.form.get('observatii', '').strip(), id))
        db.commit()
        db.close()
        flash('Predare actualizată.', 'success')
        return redirect(url_for('predari_lista'))
    predare = db.execute('SELECT * FROM predari WHERE id = ?', (id,)).fetchone()
    proiecte = db.execute("""SELECT p.id, p.nume, c.numar_contract, cl.nume as client_nume
                             FROM proiecte p
                             JOIN contracte c ON p.contract_id = c.id
                             JOIN clienti cl ON c.client_id = cl.id
                             ORDER BY p.nume""").fetchall()
    db.close()
    if not predare:
        flash('Predare negăsită.', 'danger')
        return redirect(url_for('predari_lista'))
    return render_template('predari/formular.html', predare=predare, proiecte=proiecte)


@app.route('/predari/sterge/<int:id>')
@login_required
def predari_sterge(id):
    db = get_db()
    db.execute('DELETE FROM predari WHERE id = ?', (id,))
    db.commit()
    db.close()
    flash('Predare ștearsă.', 'success')
    return redirect(url_for('predari_lista'))


# ==================== FACTURI ====================

@app.route('/facturi')
@login_required
def facturi_lista():
    db = get_db()
    cautare = request.args.get('cautare', '').strip()
    status_filtru = request.args.get('status', '').strip()
    query = """SELECT f.*, c.numar_contract, cl.nume as client_nume
               FROM facturi f
               JOIN contracte c ON f.contract_id = c.id
               JOIN clienti cl ON c.client_id = cl.id WHERE 1=1"""
    params = []
    if cautare:
        query += " AND (f.numar_factura LIKE ? OR c.numar_contract LIKE ? OR cl.nume LIKE ?)"
        params.extend([f'%{cautare}%'] * 3)
    if status_filtru:
        query += " AND f.status = ?"
        params.append(status_filtru)
    query += " ORDER BY f.data_emitere DESC"
    facturi = db.execute(query, params).fetchall()
    total_neincasat = db.execute("SELECT COALESCE(SUM(valoare), 0) as val FROM facturi WHERE status = 'Emisa'").fetchone()['val']
    db.close()
    return render_template('facturi/lista.html', facturi=facturi, cautare=cautare, status_filtru=status_filtru, total_neincasat=total_neincasat)


@app.route('/facturi/adauga', methods=['GET', 'POST'])
@login_required
def facturi_adauga():
    db = get_db()
    if request.method == 'POST':
        db.execute('''INSERT INTO facturi (numar_factura, contract_id, valoare, data_emitere, data_scadenta, status)
                      VALUES (?, ?, ?, ?, ?, ?)''',
                   (request.form['numar_factura'].strip(), request.form['contract_id'],
                    float(request.form['valoare']), request.form['data_emitere'],
                    request.form.get('data_scadenta') or None, request.form.get('status', 'Emisa')))
        db.commit()
        db.close()
        flash('Factură adăugată cu succes.', 'success')
        return redirect(url_for('facturi_lista'))
    contracte = db.execute("""SELECT c.id, c.numar_contract, cl.nume as client_nume
                              FROM contracte c JOIN clienti cl ON c.client_id = cl.id
                              ORDER BY c.numar_contract""").fetchall()
    db.close()
    return render_template('facturi/formular.html', factura=None, contracte=contracte)


@app.route('/facturi/editeaza/<int:id>', methods=['GET', 'POST'])
@login_required
def facturi_editeaza(id):
    db = get_db()
    if request.method == 'POST':
        db.execute('''UPDATE facturi SET numar_factura=?, contract_id=?, valoare=?, data_emitere=?,
                      data_scadenta=?, status=?, actualizat_la=CURRENT_TIMESTAMP WHERE id=?''',
                   (request.form['numar_factura'].strip(), request.form['contract_id'],
                    float(request.form['valoare']), request.form['data_emitere'],
                    request.form.get('data_scadenta') or None, request.form.get('status', 'Emisa'), id))
        db.commit()
        db.close()
        flash('Factură actualizată.', 'success')
        return redirect(url_for('facturi_lista'))
    factura = db.execute('SELECT * FROM facturi WHERE id = ?', (id,)).fetchone()
    contracte = db.execute("""SELECT c.id, c.numar_contract, cl.nume as client_nume
                              FROM contracte c JOIN clienti cl ON c.client_id = cl.id
                              ORDER BY c.numar_contract""").fetchall()
    db.close()
    if not factura:
        flash('Factură negăsită.', 'danger')
        return redirect(url_for('facturi_lista'))
    return render_template('facturi/formular.html', factura=factura, contracte=contracte)


@app.route('/facturi/sterge/<int:id>')
@login_required
def facturi_sterge(id):
    db = get_db()
    db.execute('DELETE FROM facturi WHERE id = ?', (id,))
    db.commit()
    db.close()
    flash('Factură ștearsă.', 'success')
    return redirect(url_for('facturi_lista'))


# ==================== PDF TO EXCEL ====================

def extract_pdf_tables(pdf_path):
    """Extrage toate tabelele din PDF folosind pdfplumber."""
    all_tables = []
    with pdfplumber.open(pdf_path) as pdf:
        for i, page in enumerate(pdf.pages):
            tables = page.extract_tables()
            for j, table in enumerate(tables):
                if table and len(table) > 0:
                    # Curăță celulele - înlocuiește None cu string gol
                    cleaned = []
                    for row in table:
                        cleaned.append([str(cell).strip() if cell else '' for cell in row])
                    all_tables.append({
                        'page': i + 1,
                        'table_index': j + 1,
                        'rows': cleaned
                    })

        # Dacă nu s-au găsit tabele, extrage textul linie cu linie
        if not all_tables:
            text_rows = []
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    for line in text.split('\n'):
                        line = line.strip()
                        if line:
                            text_rows.append([line])
            if text_rows:
                all_tables.append({
                    'page': 1,
                    'table_index': 1,
                    'rows': text_rows
                })
    return all_tables


@app.route('/pdf-to-excel', methods=['GET', 'POST'])
@login_required
def pdf_to_excel():
    if request.method == 'POST':
        if 'pdf_file' not in request.files:
            flash('Nu a fost selectat niciun fișier.', 'danger')
            return redirect(url_for('pdf_to_excel'))

        file = request.files['pdf_file']
        if file.filename == '':
            flash('Nu a fost selectat niciun fișier.', 'danger')
            return redirect(url_for('pdf_to_excel'))

        if not file.filename.lower().endswith('.pdf'):
            flash('Doar fișierele PDF sunt acceptate.', 'danger')
            return redirect(url_for('pdf_to_excel'))

        # Salvează fișierul temporar
        pdf_filename = f"temp_{current_user.id}_{file.filename}"
        pdf_path = os.path.join(UPLOAD_FOLDER, pdf_filename)
        file.save(pdf_path)

        try:
            tables = extract_pdf_tables(pdf_path)
            if not tables:
                flash('Nu s-au găsit date în fișierul PDF.', 'warning')
                os.remove(pdf_path)
                return redirect(url_for('pdf_to_excel'))

            # Salvează datele în sesiune pentru export
            session['pdf_tables'] = json.dumps(tables, ensure_ascii=False)
            session['pdf_original_name'] = file.filename

            return render_template('pdf_to_excel.html',
                                   tables=tables,
                                   original_name=file.filename,
                                   has_data=True)
        except Exception as e:
            flash(f'Eroare la procesarea PDF-ului: {str(e)}', 'danger')
            if os.path.exists(pdf_path):
                os.remove(pdf_path)
            return redirect(url_for('pdf_to_excel'))
        finally:
            if os.path.exists(pdf_path):
                os.remove(pdf_path)

    return render_template('pdf_to_excel.html', tables=None, has_data=False)


@app.route('/pdf-to-excel/download', methods=['POST'])
@login_required
def pdf_to_excel_download():
    tables_json = session.get('pdf_tables')
    original_name = session.get('pdf_original_name', 'document')

    if not tables_json:
        flash('Nu există date de exportat. Încarcă un PDF mai întâi.', 'danger')
        return redirect(url_for('pdf_to_excel'))

    tables = json.loads(tables_json)

    # Verifică ce tabele au fost selectate
    selected_tables = request.form.getlist('selected_tables')
    use_first_row_header = 'first_row_header' in request.form

    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='0D6EFD', end_color='0D6EFD', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for idx, table in enumerate(tables):
        table_key = str(idx)
        if selected_tables and table_key not in selected_tables:
            continue

        sheet_name = f"Pagina {table['page']} - Tabel {table['table_index']}"
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]
        ws = wb.create_sheet(title=sheet_name)

        rows = table['rows']
        if not rows:
            continue

        start_row = 1
        if use_first_row_header and len(rows) > 1:
            for col_idx, header_val in enumerate(rows[0], 1):
                cell = ws.cell(row=1, column=col_idx, value=header_val)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            start_row = 2
            rows = rows[1:]

        for row_idx, row in enumerate(rows, start_row):
            for col_idx, val in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = cell_alignment
                cell.border = thin_border

        # Auto-dimensionare coloane
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 4, 50)

    if len(wb.sheetnames) == 0:
        flash('Nu au fost selectate tabele pentru export.', 'warning')
        return redirect(url_for('pdf_to_excel'))

    # Salvează Excel temporar și trimite-l
    excel_name = os.path.splitext(original_name)[0] + '.xlsx'
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx', dir=UPLOAD_FOLDER)
    wb.save(tmp.name)
    tmp.close()

    response = send_file(
        tmp.name,
        as_attachment=True,
        download_name=excel_name,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    # Cleanup after send
    @response.call_on_close
    def cleanup():
        if os.path.exists(tmp.name):
            os.remove(tmp.name)

    return response


# Initialize DB and admin on import (needed for gunicorn)
init_db()
create_admin_if_needed()

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)), debug=True)
